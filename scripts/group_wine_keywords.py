import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_excel('wine-site-seo-keywords.xlsx', sheet_name='Best Wine For')
keywords = df['Keyword'].dropna().tolist()
terms = sorted(set(k.strip() for k in keywords))

THEMES = [
    'Beef Dishes',
    'Pasta & Italian Dishes',
    'Risotto',
    'Seafood & Fish',
    'Poultry (Chicken, Turkey, Duck)',
    'Pork & Charcuterie',
    'Lamb & Game',
    'Soups, Stews & Braises',
    'BBQ & Grilled Food',
    'Vegetarian & Vegetables',
    'Cheese',
    'Fondue & Sharing',
    'Desserts & Chocolate',
    'Asian & International Cuisine',
    'Indian Food',
    'Mexican & Latin Food',
    'Other International Cuisines',
    'Cooking Wine (General)',
    'Sangria & Cocktail Mixing',
    'Mulled Wine & Hot Spiced Wine',
    'Mimosa & Brunch Drinks',
    'Gifts & Occasions',
    'Beginners & Casual Drinking',
    'Health & Dietary Concerns',
    'Season & Weather',
    'Pizza & Burgers',
    'Other / Miscellaneous',
]

themes = {t: [] for t in THEMES}


def classify(kw):
    k = kw.lower()

    # Risotto (before pasta to avoid overlap)
    if 'risotto' in k:
        return 'Risotto'

    # Mulled wine and hot spiced wine
    if any(x in k for x in ['mulled', 'gluhwein', 'glühwein', 'glühwein', 'glogg', 'gløgg',
                              'hot spiced wine', 'vin chaud', 'mulled wine', 'vin brule',
                              'vin brulé', 'wassail', 'hot wine']):
        return 'Mulled Wine & Hot Spiced Wine'

    # Sangria & cocktail mixing
    if any(x in k for x in ['sangria', 'aperol', 'cocktail', 'punch', 'apple cider', 'spritz',
                              'mimosa mix', 'mixer', 'kir royale', ' kir ', 'kir wine', 'for kir',
                              'kalimotxo', 'tinto de verano', 'new york sour', 'ninja slushie',
                              'wine spritzer', 'wine cooler']):
        if 'mimosa' not in k:
            return 'Sangria & Cocktail Mixing'

    # Mimosa & brunch
    if 'mimosa' in k or 'brunch' in k:
        return 'Mimosa & Brunch Drinks'

    # Pasta & Italian (broad)
    pasta_italian = ['pasta', 'spaghetti', 'bolognese', 'ragu', 'lasagna', 'lasagne',
                     'fettuccine', 'linguine', 'penne', 'ravioli', 'gnocchi', 'rigatoni',
                     'tagliatelle', 'carbonara', 'arrabbiata', 'arrabiata', 'amatriciana',
                     'cacio e pepe', 'puttanesca', 'aglio e olio', 'baked ziti', 'manicotti',
                     'alfredo', 'osso buco', 'polenta', 'calzone', 'focaccia', 'bruschetta',
                     'caprese', 'saltimbocca', 'zabaglione', 'tiramisu', 'panna cotta',
                     'zuppa toscana', 'marsala', 'piccata', 'eggplant parmesan',
                     'eggplant parmigiana', 'baked pasta', 'cannelloni', 'marinara',
                     'spag bol', 'italian food', 'italian dinner', 'italian meal',
                     ' ziti', 'vongole', 'for italian']
    if any(x in k for x in pasta_italian):
        if any(x in k for x in ['tiramisu', 'panna cotta', 'zabaglione', 'dessert']):
            return 'Desserts & Chocolate'
        return 'Pasta & Italian Dishes'

    # Pizza & burgers
    if any(x in k for x in ['pizza', 'burger', 'hamburger', 'hot dog', 'sandwich', 'kebab',
                              'cheeseburger']):
        return 'Pizza & Burgers'

    # Beef
    beef_terms = ['beef', 'steak', 'brisket', 'tenderloin', 'ribeye', 'rib eye', 'wellington',
                  'stroganoff', 'chuck roast', 'pot roast', 'short rib', 'boeuf', 'bourguignon',
                  'prime rib', 'carpaccio', 'meatball', 'meatloaf', 'au jus', 'beef fajita',
                  'beef taco', 'beef enchilada', 'filet mignon', 'eye fillet', 'flank steak',
                  't-bone', 'sirloin', 'rib roast', 'beef roast', 'beef dinner', 'beef dishes',
                  'beef ragu', 'beef stew', 'beef curry', 'beef chili', 'beef casserole',
                  'beef pasta', 'beef gravy', 'beef jus', 'beef marinade', 'beef stock',
                  'beef fondue', 'braised short', 'beef short', 'beef ribs', 'beef cheeks',
                  'beef medallion', 'beef wellington', 'beef kabob', 'beef osso', 'beef pho',
                  'beef pie', 'beef rendang', 'beef tacos', 'beef tips', 'beef tenderloin',
                  'cottage pie', 'shepherds pie', "shepherd's pie", 'beef daube', 'beef fillet',
                  'beef barley', 'beef burgundy', 'beef bergen', 'veal', 'oxtail',
                  'scotch fillet', 'new york strip', 'wagyu']
    if any(x in k for x in beef_terms):
        if 'burger' in k:
            return 'Pizza & Burgers'
        return 'Beef Dishes'

    # Pork
    pork_terms = ['pork', 'bacon', ' ham', 'prosciutto', 'salami', 'chorizo', 'sausage',
                  'charcuterie', 'pulled pork', 'mortadella', 'pancetta', 'guanciale',
                  'easter ham', 'ham dinner', 'serrano', 'iberico', 'bratwurst', 'hot dog',
                  'pork belly', 'pork rib', 'pork chop', 'pork loin', 'pork roast', 'pork fillet',
                  'pig', 'suckling', 'porchetta', 'jamon', 'jambon', 'coppa', 'lardons']
    if any(x in k for x in pork_terms):
        return 'Pork & Charcuterie'

    # Poultry
    poultry_terms = ['chicken', 'turkey', 'duck', 'goose', 'poultry', 'hen', 'coq au vin',
                     'schnitzel', 'roast chicken', 'roast turkey', 'roast duck', 'chicken curry',
                     'chicken tikka', 'grilled chicken', 'rotisserie', 'roast fowl']
    if any(x in k for x in poultry_terms):
        return 'Poultry (Chicken, Turkey, Duck)'

    # Seafood
    seafood_terms = ['fish', 'salmon', 'tuna', 'halibut', 'cod', 'sea bass', 'trout', 'tilapia',
                     'snapper', 'swordfish', 'barramundi', 'crab', 'lobster', 'shrimp', 'prawn',
                     'scallop', 'oyster', 'mussel', 'clam', 'seafood', 'sushi', 'ceviche',
                     'paella', 'bouillabaisse', 'calamari', 'squid', 'octopus', 'anchov',
                     'escargot', 'fish and chips', 'fish pie', 'garlic prawns', 'fish tacos',
                     'grilled fish', 'battered fish', 'fishcake', 'fish cake']
    if any(x in k for x in seafood_terms):
        return 'Seafood & Fish'

    # Lamb & game
    game_terms = ['lamb', 'mutton', 'venison', 'rabbit', 'bison', 'wild boar', 'elk',
                  'game meat', 'haggis', 'moose', 'deer', 'pheasant', 'quail', 'partridge']
    if any(x in k for x in game_terms):
        return 'Lamb & Game'

    # Indian food
    indian_terms = ['indian', 'tikka', 'korma', 'vindaloo', 'masala', 'dal', 'dhal',
                    'paneer', 'biryani', 'naan', 'samosa', 'chutney', 'curry leaf',
                    'saag', 'palak', 'butter chicken', 'tandoori', 'dosa', 'idli',
                    'for india', 'for curry']
    if any(x in k for x in indian_terms):
        return 'Indian Food'

    # Asian & international (broad)
    asian_terms = ['asian', 'chinese', 'japanese', 'korean', 'thai', 'vietnamese', 'cantonese',
                   'dim sum', 'pad thai', 'teriyaki', 'stir fry', 'stir-fry', 'noodle',
                   'dumpling', 'wonton', 'bibimbap', 'kimchi', 'bao bun', 'tempura',
                   'fried rice', 'satay', 'rendang', 'banh mi', 'ramen', ' pho',
                   'hot pot', 'hotpot', 'peking duck', 'szechuan', 'taiwanese', 'malay',
                   'singapore', 'indonesian', 'burmese', 'cambodian', 'lao', 'green curry',
                   'red curry', 'massaman', 'dumplings', 'gyoza', 'miso', 'yakitori',
                   'bulgogi', 'tteokbokki', 'katsu', 'laksa', 'omakase', 'yellow curry',
                   'jamaican', 'jambalaya']
    if any(x in k for x in asian_terms):
        return 'Asian & International Cuisine'

    # Mexican & Latin
    latin_terms = ['mexican', 'taco', 'burrito', 'enchilada', 'quesadilla', 'fajita',
                   'guacamole', 'salsa', 'nacho', 'tamale', 'carnitas', 'mole', 'pozole',
                   'latin', 'peruvian', 'argentinian', 'brazilian', 'spanish food',
                   'tapas', 'empanada', 'ceviche', 'chimichurri']
    if any(x in k for x in latin_terms):
        return 'Mexican & Latin Food'

    # Other international
    intl_terms = ['greek', 'ethiopian', 'moroccan', 'lebanese', 'middle eastern', 'turkish',
                  'french food', 'german food', 'polish food', 'russian food', 'african',
                  'mediterranean', 'mediterranean food', 'haggis', 'scottish', 'irish stew',
                  'gumbo', 'cajun']
    if any(x in k for x in intl_terms):
        return 'Other International Cuisines'

    # BBQ & grilled
    if any(x in k for x in ['bbq', 'barbecue', 'grill', 'grilled', 'smoked meat', 'smoke pit',
                              'smokehouse']):
        return 'BBQ & Grilled Food'

    # Soups, stews, braises
    stew_terms = ['soup', 'stew', 'braise', 'casserole', 'chowder', 'bisque', 'broth',
                  'goulash', 'tagine', 'daube', 'braised', 'chili', 'slow cooker',
                  'crockpot', 'french onion', 'minestrone', 'bouillabaisse', 'potage',
                  'gumbo', 'pot au feu', 'hot pot', 'hotpot', 'beef barley', 'lentil soup',
                  'tomato soup', 'mushroom soup', 'clam chowder']
    if any(x in k for x in stew_terms):
        return 'Soups, Stews & Braises'

    # Cheese
    cheese_terms = ['cheese', 'brie', 'camembert', 'cheddar', 'gouda', 'parmesan',
                    'mozzarella', 'feta', 'gorgonzola', 'roquefort', 'manchego', 'gruyere',
                    'stilton', 'blue cheese', 'cheeseboard', 'cheese board', 'ricotta',
                    'pecorino', 'comte', 'emmental', 'swiss cheese', 'aged cheese']
    if any(x in k for x in cheese_terms):
        return 'Cheese'

    # Fondue & raclette
    if 'fondue' in k or 'raclette' in k:
        return 'Fondue & Sharing'

    # Desserts & chocolate
    dessert_terms = ['chocolate', 'dessert', 'cake', 'cheesecake', 'pavlova', 'fruit cake',
                     'brownie', 'tart', 'pastry', 'ice cream', 'pudding', 'mousse',
                     'profiterole', 'creme brulee', 'tiramisu', 'panna cotta', 'zabaglione',
                     'biscotti', 'cannoli', 'eclair', 'macaroon', 'macaron', 'crepe',
                     'waffle', 'pancake dessert', 'caramel', 'trifle', 'key lime pie']
    if any(x in k for x in dessert_terms):
        return 'Desserts & Chocolate'

    # Vegetarian & vegetables
    veg_terms = ['vegetarian', 'vegan', 'vegetable', 'mushroom', 'asparagus', 'artichoke',
                 'zucchini', 'courgette', 'salad', 'lentil', 'tofu', 'tempeh', 'cauliflower',
                 'broccoli', 'spinach', 'kale', 'butternut', 'pumpkin', 'squash', 'beetroot',
                 'caprese', 'hummus', 'falafel', 'ratatouille', 'quiche', 'souffle',
                 'tabbouleh', 'tzatziki', 'veggie', 'portobello', 'roast veg',
                 'stuffed peppers', 'sweet potato', 'corn', 'eggplant', 'aubergine',
                 'nut roast', 'jalapeno', 'hot wings']
    if any(x in k for x in veg_terms):
        return 'Vegetarian & Vegetables'

    # Cooking wine general
    cooking_terms = ['cooking', 'baking', 'deglazing', 'demi glace', 'marinad', 'braising wine',
                     'baking cake', 'baking turkey', 'baking fruit', 'fruit cake', 'gravy',
                     'jus ', 'for jus', 'red wine jus', 'mulling spices', 'stock', 'sauce',
                     'recipe wine', 'for cooking', 'for baking', 'slow cook']
    if any(x in k for x in cooking_terms):
        if not any(x in k for x in ['beef', 'chicken', 'lamb', 'pork', 'risotto', 'pasta',
                                     'seafood', 'fish', 'steak', 'turkey', 'duck', 'rabbit',
                                     'vegetable']):
            return 'Cooking Wine (General)'

    # Gifts & occasions
    occasion_terms = ['gift', 'occasion', 'birthday', 'anniversary', 'wedding', 'christmas',
                      'xmas', 'holiday', 'valentines', "valentine's", 'thanksgiving', 'mother',
                      'father', 'graduation', 'new year', 'housewarming', 'retirement', 'party',
                      'celebration', 'date night', 'romantic', 'easter', 'passover', 'diwali',
                      'corporate', 'boss', 'colleague', 'hostess', 'host gift', 'hamper',
                      'basket', 'picnic', 'special occasion', 'nye', 'present', 'zodiac']
    if any(x in k for x in occasion_terms):
        return 'Gifts & Occasions'

    # Beginners & casual
    casual_terms = ['beginner', 'first time', 'first timer', 'casual', 'everyday', 'everyday drinking',
                    'easy drinking', 'non-wine', 'wine lover', 'wine drinker', 'someone who',
                    'woman', 'women', 'female', 'girl', 'lady', 'ladies', 'men', 'guys',
                    'cheap wine', 'budget wine', 'affordable', 'starter', 'newbie', 'wine newbie',
                    'entertaining', 'dinner party', 'for drinking', 'evening drinking',
                    'afternoon', 'elderly', 'drinking wine', 'wine drinking', 'wine night',
                    'happy hour', 'house wine', 'daily drink', 'daily wine', 'old people',
                    "don't like wine", 'just drinking', 'for dinner', 'for evening', 'for you',
                    'young adults', 'the money', 'your buck', 'me quiz', 'quiz', 'nasha',
                    'under $', 'under 10', 'under £', 'under €', 'under 20', 'under 30',
                    'under 40', 'under 50', 'under 100', 'under 200', 'good wine for under']
    if any(x in k for x in casual_terms):
        return 'Beginners & Casual Drinking'

    # Health & dietary
    health_terms = ['acid reflux', 'low acid', 'low sugar', 'low calorie', 'low carb',
                    'diabetic', 'diabetes', 'health', 'keto', 'gluten', 'hangover', 'headache',
                    'sulfite', 'organic wine', 'natural wine', 'anxiety', 'allerg',
                    'heart', 'blood pressure', 'cholesterol', 'gout', 'gerd', 'reflux',
                    'digestion', 'gut health', 'skin', 'endometriosis', 'energy', 'immune',
                    'diet wine', 'weight loss', 'anti-inflammatory', 'antioxidant', 'resveratrol',
                    'low tannin', 'low histamine', 'headache free', 'ibs', 'inflammation',
                    'insulin', 'iron deficiency', 'kidney', 'migraine', 'ulcer',
                    'upset stomach', 'uric acid', 'yeast intolerance', 'period',
                    'your stomach', 'zero alcohol', 'weight gain', 'for diet']
    if any(x in k for x in health_terms):
        return 'Health & Dietary Concerns'

    # Season & weather
    season_terms = ['summer', 'winter', 'spring', 'autumn', 'fall ', 'hot day', 'hot weather',
                    'cold weather', 'beach', 'picnic', 'outdoor', 'season', 'al fresco',
                    'chilled red', 'warm weather', 'cool weather', 'winter warmer', 'hot tub']
    if any(x in k for x in season_terms):
        return 'Season & Weather'

    return 'Other / Miscellaneous'


for kw in terms:
    theme = classify(kw)
    themes[theme].append(kw)

print('=== THEME COUNTS ===')
for theme in THEMES:
    count = len(themes[theme])
    print(f'{count:4d}  {theme}')
total = sum(len(v) for v in themes.values())
print(f'\nTotal placed: {total}')
print(f'Input total:  {len(terms)}')

print('\n=== REMAINING MISC ===')
for kw in themes['Other / Miscellaneous']:
    print(' ', kw)

# Build output dataframe: theme column + keyword column
rows = []
for theme in THEMES:
    for kw in sorted(themes[theme]):
        rows.append({'Theme': theme, 'Keyword': kw})
out_df = pd.DataFrame(rows)

# Write to new sheet in existing workbook
wb = load_workbook('wine-site-seo-keywords.xlsx')

sheet_name = 'Best Wine For - Grouped'
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

# Theme color palette (one per theme, cycling soft pastels)
COLORS = [
    'FFDDC1', 'FFE4B5', 'FFF9C4', 'D4EDDA', 'CCE5FF', 'E2D9F3',
    'F8D7DA', 'D1ECF1', 'FFF3CD', 'F0E6FF', 'E8F5E9', 'FCE4EC',
    'E3F2FD', 'FFF8E1', 'F3E5F5', 'E0F2F1', 'FBE9E7', 'EDE7F6',
    'E8EAF6', 'F9FBE7', 'FFF3E0', 'E8F5E9', 'F3E5F5', 'E1F5FE',
    'F9FBE7', 'EEEEEE', 'F5F5F5',
]

theme_color = {theme: COLORS[i % len(COLORS)] for i, theme in enumerate(THEMES)}

# Headers
headers = ['Theme', 'Keyword']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    cell.fill = PatternFill('solid', start_color='2F4F4F')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.row_dimensions[1].height = 22

# Data rows
for row_idx, row in enumerate(rows, 2):
    theme = row['Theme']
    color = theme_color[theme]
    for col_idx, field in enumerate(['Theme', 'Keyword'], 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=row[field])
        cell.font = Font(name='Arial', size=10)
        cell.fill = PatternFill('solid', start_color=color)
        cell.alignment = Alignment(vertical='center', wrap_text=False)

# Column widths
ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 52

# Freeze header
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = f'A1:B{len(rows) + 1}'

wb.save('wine-site-seo-keywords.xlsx')
print(f'\nSaved {len(rows)} rows to sheet "{sheet_name}"')
